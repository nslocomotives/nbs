from openpyxl import Workbook
from openpyxl.styles import Font, Fill
import database
import datetime

wb = Workbook()

months = {'Jan':'01', 'Feb':'02', 'Mar':'03', 'Apr':'04','May':'05', 'Jun':'06', 'Jul':'07', 'Aug':'08', 'Sep':'09', 'Oct':'10', 'Nov':'11', 'Dec':'12'}
incomeCats=database.getCategory(1)
expensesCats=database.getCategory(0)
subcat={}
for i in incomeCats:
    subcat[i['name']]=database.getSubCategory(i['id'])
for e in expensesCats:
    subcat[e['name']]=database.getSubCategory(e['id'])

# Get all the account names and the releveant info to get the data.
accounts=database.getAccounts()


houseExpenses=['Lodging Rent',
               'Lodging',
               'Lodging Tax (council)',
               'Lodging Utitlities',
               'Lodging Maintence',
               'Lodging Insurance',
               'Lodging Other']
djExpenses=['DJ Software',
            'DJ Insurance',
            'DJ Equipment',
            'DJ Other']
personalExpenses=['Shopping',
                  'Clothing',
                  'Entertainment',
                  'Crypto Investment',
                  'Dating',
                  'Eating/drinking Out',
                  'Birthdays and Christmas',
                  'Car Maintenance',
                  'Fuel',
                  'Child Maintence',
                  'FD Loan',
                  'Hitachi Loan',
                  'Mobile Phone',
                  'Life Insurance',
                  'TV License',
                  'Bank Charges',
                  'Debt repayment',
                  'Other Expenses']
expenses=houseExpenses+djExpenses+personalExpenses

# grab the active worksheet and create other worksheets and give them names
wsO = wb.active
wsO.title = "Overview"
wsB = wb.create_sheet("Budget")

#set starting rows
wsOrow = 2
wsBrow = 2
wsMrow = 2

#function to wrap a get account names and batch up the results
def getAccountsList(AccTypes):
    MergedList = []
    #DEBUG print('Content of MergedList', MergedList)
    for a in AccTypes:
        List = database.getAccountNames(a)
        #DEBUG print('Fetched from database:', List)
        MergedList.extend(List)
        #DEBUG print('Content of MergedList', MergedList)
    return(MergedList)

#function to wrap a get balances of accounts and clean up results 
def getBalance(accountnumber, Date):
    result = database.getBal(accountnumber, Date)
    #DEBUG print('Content of result', type(result), result)
    if type(result) == list:
        result = result[0]
    else:
        result = {'balance':0}
    #DEBUG print('Content of result', type(result), result)
    return(result)
        

#create the budget header
wsB['A' + str(wsBrow)] = 'Monthly Expenses'
wsB['E' + str(wsBrow)] = 'monthly Income'
wsBHrow = wsBrow
wsBrowExp = wsBrow + 1
wsBrowInc = wsBrow + 1
#create the budget rows
for e in expensesCats:
    for f in subcat[e['name']]:
        wsB['A' + str(wsBrowExp)] = f['name']
        wsBrowExp = wsBrowExp + 1
for i in incomeCats:
    for j in subcat[i['name']]: 
        wsB['E' + str(wsBrowInc)] = j['name']
        wsBrowInc = wsBrowInc + 1
#create the budget totals
wsB['A' + str(wsBrowExp)] = 'Total Expenses'
wsB['B' + str(wsBrowExp)] = '=SUM(B' + str(wsBHrow + 1) + ':B' + str(wsBrowExp - 1) +')'
BExpTotal = 'B' + str(wsBrowExp)
wsB['E' + str(wsBrowInc)] = 'Total Income'
wsB['F' + str(wsBrowInc)] = '=SUM(F' + str(wsBHrow + 1) + ':F' + str(wsBrowInc - 1) +')'
BIncTotal = 'F' + str(wsBrowInc)

#create column headers


#create the month.
def create_month(ws):
    mmm = ws
    ws = wb.create_sheet(ws)
    wsMrowInc = wsMrow #starting row for Income Column
    wsMrowAna = wsMrow #starting row for Analysis Column
    wsMrowData = wsMrow #starting row for Bank Data
    column = 'A' #Starting column for Income and Analysis 
    columnData = 'I' #starting column for Bank Data
    IncTotal ={}
    ExpTotal ={}
    AccColumns = {}
    Balance = {}
    #set up some date strings...
    #todays date when the script is run
    now = datetime.datetime.now()
    #change month from mmm to mm format
    month = months[mmm]
    
    #add this year when the script is run to the month and build a start and end date for use later
    startDate = str(now.year) + '-' + str(month) + '-01'
    endDate = str(now.year) + '-' + str(month) + '-31'

    #### PRODUCE DATA TABLES FROM BANK DATA ####
    #DEBUG print(accounts)
    for a in accounts:
        #get transaction data
        Balance[a['accountname']] = {}
        #set opening and closing balances
        Balance[a['accountname']]['open'] = getBalance(a['accountnumber'], startDate)
        Balance[a['accountname']]['close'] = getBalance(a['accountnumber'], endDate)
        print(a['accountname'], a['accountnumber'],startDate, endDate)
        print(Balance[a['accountname']])
        transactions=database.getTransactionDataByAccount(a['accountnumber'], startDate, endDate)
        #write title of account
        ws[columnData + str(wsMrowData)] = a['accountname']
        ws[columnData + str(wsMrowData)].font = Font(bold=True)
        ws.merge_cells(columnData + str(wsMrowData) + ':' + chr(ord(columnData)+2) + str(wsMrowData))
        wsMrowData = wsMrowData + 1
        #set the columns to add to the totals columns
        AccColumns[a['accountname']] = {}
        AccColumns[a['accountname']]['Date'] = columnData
        AccColumns[a['accountname']]['Description'] = chr(ord(columnData)+1)
        AccColumns[a['accountname']]['Category'] = chr(ord(columnData)+2)
        AccColumns[a['accountname']]['Outgoings'] = chr(ord(columnData)+3)
        AccColumns[a['accountname']]['Income'] = chr(ord(columnData)+4)
        #DEBUG print (AccColumns)
        ws[columnData + str(wsMrowData)] = 'Date'
        ws[AccColumns[a['accountname']]['Date'] + str(wsMrowData)] = 'Details'
        ws[AccColumns[a['accountname']]['Category'] + str(wsMrowData)] = 'Category'
        ws[AccColumns[a['accountname']]['Outgoings'] + str(wsMrowData)] = 'Outgoings'
        ws[AccColumns[a['accountname']]['Income'] + str(wsMrowData)] = 'Income'
        wsMrowData = wsMrowData + 1
        for t in transactions:
            ws[AccColumns[a['accountname']]['Date'] + str(wsMrowData)] = t['date']
            ws[AccColumns[a['accountname']]['Description'] + str(wsMrowData)] = t['description']
            if t['subcategory_lnk']:
                subCatLnk = database.getSubCategoryById(t['subcategory_lnk'])
                #DEBUG print(subCatLnk[0]['name'])
                ws[AccColumns[a['accountname']]['Category'] + str(wsMrowData)] = subCatLnk[0]['name']
            if t['amount'] > 0:
                ws[AccColumns[a['accountname']]['Income'] + str(wsMrowData)] = abs(t['amount'])
            else:
                ws[AccColumns[a['accountname']]['Outgoings'] + str(wsMrowData)] = abs(t['amount'])
            wsMrowData = wsMrowData + 1
        #set next start column for the next Account details
        columnData = chr(ord(columnData)+6)
        #reset the row counter
        wsMrowData = wsMrow

    
    #### PRODUCE TOTALS COLUMNS ####
    #income column
    # header
    ws[column + str(wsMrowInc)] = 'Income'
    ws[column + str(wsMrowInc)].font = Font(bold=True)
    ws.merge_cells(column + str(wsMrowAna) + ':' + chr(ord(column)+2) + str(wsMrowAna))
    wsMrowInc = wsMrowInc + 1
    #income section
    for i in incomeCats:
      ws[column + str(wsMrowInc)] = i['name'] + ' Income'
      ws[column + str(wsMrowInc)].font = Font(bold=True)
      wsMHrowInc = wsMrowInc
      wsMrowInc = wsMrowInc + 1
      ##income catagories
      #print(subcat[i['name']])
      for j in subcat[i['name']]:
        ws[chr(ord(column)+1) + str(wsMrowInc)] = j['name']
        catSum = '=SUM('
        #DEBUG print(AccColumns)
        for k, v in AccColumns.items():
            #DEBUG print(v)
            catC = v['Category']
            incC = v['Income']
            outC = v['Outgoings']
            comC = chr(ord(column)+1) + str(wsMrowInc)
            catSum = catSum + 'SUMIF($' + catC + ':$' + catC + ',' + comC + ',$' + incC + ':$' + incC + '),'
            catSum = catSum + 'SUMIF($' + catC + ':$' + catC + ',' + comC + ',$' + outC + ':$' + outC + '),'
        catSum = catSum[:-1] + ')'
        #DEBUG print(catSum)
        ws[chr(ord(column)+2) + str(wsMrowInc)] = str(catSum)  
        wsMrowInc = wsMrowInc + 1
      ##income total
      ws[chr(ord(column)+1) + str(wsMrowInc)] = i['name'] + ' Total'
      ws[chr(ord(column)+1) + str(wsMrowInc)].font = Font(bold=True)
      ws[chr(ord(column)+2) + str(wsMrowInc)] = '=SUM(' + chr(ord(column)+2) + str(wsMHrowInc + 1) + ':' + chr(ord(column)+2) + str(wsMrowInc - 1) +')'
      ws[chr(ord(column)+2) + str(wsMrowInc)].font = Font(bold=True)
      IncTotal[i['name']] = chr(ord(column)+2) + str(wsMrowInc)
      #print(IncTotal[i['name']])
      wsMrowInc = wsMrowInc + 1
      
    ##Total Income
    ws[column + str(wsMrowInc)] = 'Total Income'
    ws[column + str(wsMrowInc)].font = Font(bold=True)
    concatStr = '=SUM('
    #print(IncTotal)
    for k in IncTotal:
        concatStr = concatStr + IncTotal[k] + '+'
    ws[chr(ord(column)+2) + str(wsMrowInc)] = concatStr[:-1] + ')'
    ws[chr(ord(column)+2) + str(wsMrowInc)].font = Font(bold=True)
    IncTotal['Total'] = chr(ord(column)+2) + str(wsMrowInc)
    #print(IncTotal)
    wsMrowInc = wsMrowInc + 2

    #Expenses column
    # header
    ws[column + str(wsMrowInc)] = 'Expenses'
    ws[column + str(wsMrowInc)].font = Font(bold=True)
    ws.merge_cells(column + str(wsMrowAna) + ':' + chr(ord(column)+2) + str(wsMrowAna))
    wsMrowInc = wsMrowInc + 1

    #expenses section
    for e in expensesCats:
      ws[column + str(wsMrowInc)] = e['name'] + ' Expenses'
      ws[column + str(wsMrowInc)].font = Font(bold=True)
      wsMHrowInc = wsMrowInc
      wsMrowInc = wsMrowInc + 1
      ##expenses catagories
      for f in subcat[e['name']]:
        ws[chr(ord(column)+1) + str(wsMrowInc)] = f['name']
        if f['name'] == 'Tax(Income)':
            ExpTotal['TaxIncome'] = chr(ord(column)+2) + str(wsMrowInc)
        catSum = '=SUM('
        #DEBUG print(AccColumns)
        for k, v in AccColumns.items():
            #DEBUG print(v)
            catC = v['Category']
            incC = v['Income']
            outC = v['Outgoings']
            comC = chr(ord(column)+1) + str(wsMrowInc)
            catSum = catSum + 'SUMIF($' + catC + ':$' + catC + ',' + comC + ',$' + incC + ':$' + incC + '),'
            catSum = catSum + 'SUMIF($' + catC + ':$' + catC + ',' + comC + ',$' + outC + ':$' + outC + '),'
        catSum = catSum[:-1] + ')'
        #DEBUG print(catSum)
        ws[chr(ord(column)+2) + str(wsMrowInc)] = str(catSum)
        wsMrowInc = wsMrowInc + 1
      ##expenses total
      ws[chr(ord(column)+1) + str(wsMrowInc)] = e['name'] + ' Total'
      ws[chr(ord(column)+1) + str(wsMrowInc)].font = Font(bold=True)
      ws[chr(ord(column)+2) + str(wsMrowInc)] = '=SUM(' + chr(ord(column)+2) + str(wsMHrowInc + 1) + ':' + chr(ord(column)+2) + str(wsMrowInc - 1) +')'
      ws[chr(ord(column)+2) + str(wsMrowInc)].font = Font(bold=True)
      ExpTotal[e['name']] = chr(ord(column)+2) + str(wsMrowInc)
      #print(IncTotal[e['name']])
      wsMrowInc = wsMrowInc + 1 

    ##Total Expenses
    ws[column + str(wsMrowInc)] = 'Total Expenses'
    ws[column + str(wsMrowInc)].font = Font(bold=True)
    concatStr = '=SUM('
    for g in ExpTotal:
        concatStr = concatStr + ExpTotal[g] + '+'
    ws[chr(ord(column)+2) + str(wsMrowInc)] = concatStr[:-1] + ')'
    ws[chr(ord(column)+2) + str(wsMrowInc)].font = Font(bold=True)
    ExpTotal['Total'] = chr(ord(column)+2) + str(wsMrowInc)
    wsMrowInc = wsMrowInc + 2

    #Net monthly Cash flow
    ws[column + str(wsMrowInc)] = 'Net Monthly Cash Flow'
    ws[column + str(wsMrowInc)].font = Font(bold=True)
    ws[chr(ord(column)+2) + str(wsMrowInc)] = '=SUM(' + IncTotal['Total'] + '-' + ExpTotal['Total'] + ')'
    ws[chr(ord(column)+2) + str(wsMrowInc)].font = Font(bold=True)
    NetTotal = chr(ord(column)+2) + str(wsMrowInc)
    wsMrowInc = wsMrowInc + 2

    #set liabilities start row for use later
    wsMrowLib = wsMrowInc

    #Assets Totals
    # header
    AssetTotal = {}
    ws[column + str(wsMrowInc)] = 'Assets'
    ws[column + str(wsMrowInc)].font = Font(bold=True)
    ws.merge_cells(column + str(wsMrowAna) + ':' + chr(ord(column)+2) + str(wsMrowAna))
    wsMHrowInc = wsMrowInc
    wsMrowInc = wsMrowInc + 1
    # asset totals - bank
    ws[chr(ord(column)+1) + str(wsMrowInc)] = 'Bank Accounts'
    AssetTotal['Bank Accounts'] = chr(ord(column)+2) + str(wsMrowInc)
    wsMrowInc = wsMrowInc + 1
    # asset totals - Crypto Investments
    ws[chr(ord(column)+1) + str(wsMrowInc)] = 'Crypto Investments'
    AssetTotal['Crypto Investments'] = chr(ord(column)+2) + str(wsMrowInc)
    wsMrowInc = wsMrowInc + 1
    # asset totals - Stocks/Shares
    ws[chr(ord(column)+1) + str(wsMrowInc)] = 'Stocks/Shares'
    AssetTotal['Stocks/Shares'] = chr(ord(column)+2) + str(wsMrowInc)
    wsMrowInc = wsMrowInc + 1
    # asset totals - Bonds
    ws[chr(ord(column)+1) + str(wsMrowInc)] = 'Bonds'
    AssetTotal['Bonds'] = chr(ord(column)+2) + str(wsMrowInc)
    wsMrowInc = wsMrowInc + 1
    # asset totals - Buy to let Property
    ws[chr(ord(column)+1) + str(wsMrowInc)] = 'Buy to let Property'
    AssetTotal['Buy to let Property'] = chr(ord(column)+2) + str(wsMrowInc)
    wsMrowInc = wsMrowInc + 1
    ws[chr(ord(column)+1) + str(wsMrowAna)] = '(fair market value less mortgage)'
    ws[chr(ord(column)+1) + str(wsMrowAna)].font = Font(size=8)
    #assets total
    ws[column + str(wsMrowInc)] = 'Assets Total'
    ws[column + str(wsMrowInc)].font = Font(bold=True)
    ws.merge_cells(column + str(wsMrowAna) + ':' + chr(ord(column)+1) + str(wsMrowAna))
    ws[chr(ord(column)+2) + str(wsMrowInc)] = '=SUM(' + chr(ord(column)+2) + str(wsMHrowInc + 1) + ':' + chr(ord(column)+2) + str(wsMrowInc - 1) +')'
    ws[chr(ord(column)+2) + str(wsMrowInc)].font = Font(bold=True)
    AssetTotal['Total'] = chr(ord(column)+2) + str(wsMrowInc)
    wsMrowInc = wsMrowInc + 2


    #dep Assets Totals
    # header
    DepAssetTotal = {}
    ws[column + str(wsMrowInc)] = 'Depreciting Assets'
    ws[column + str(wsMrowInc)].font = Font(bold=True)
    ws.merge_cells(column + str(wsMrowAna) + ':' + chr(ord(column)+2) + str(wsMrowAna))
    wsMHrowInc = wsMrowInc
    wsMrowInc = wsMrowInc + 1
    # get a list and value of depriciating assets from database for this month
    DepAssets = database.getAssets(endDate)
    for d in DepAssets:
        ws[chr(ord(column)+1) + str(wsMrowInc)] = d['name']
        DepAssetTotal[d['name']] = chr(ord(column)+2) + str(wsMrowInc)
        ws[DepAssetTotal[d['name']]] = d['value']
        wsMrowInc = wsMrowInc + 1
    #assets total
    ws[column + str(wsMrowInc)] = 'Depreciting Assets Total'
    ws[column + str(wsMrowInc)].font = Font(bold=True)
    ws.merge_cells(column + str(wsMrowAna) + ':' + chr(ord(column)+1) + str(wsMrowAna))
    ws[chr(ord(column)+2) + str(wsMrowInc)] = '=SUM(' + chr(ord(column)+2) + str(wsMHrowInc + 1) + ':' + chr(ord(column)+2) + str(wsMrowInc - 1) +')'
    ws[chr(ord(column)+2) + str(wsMrowInc)].font = Font(bold=True)
    DepAssetTotal['Total'] = chr(ord(column)+2) + str(wsMrowInc)
    wsMrowInc = wsMrowInc + 2

    #assets total per banker
    ws[column + str(wsMrowInc)] = 'Total Assets Per Banker'
    ws[column + str(wsMrowInc)].font = Font(bold=True)
    ws.merge_cells(column + str(wsMrowAna) + ':' + chr(ord(column)+1) + str(wsMrowAna))
    ws[chr(ord(column)+2) + str(wsMrowInc)] = '=SUM(' + AssetTotal['Total'] + '+' + DepAssetTotal['Total'] +')'
    ws[chr(ord(column)+2) + str(wsMrowInc)].font = Font(bold=True)
    AssetTotal['TotalPerBanker'] = chr(ord(column)+2) + str(wsMrowInc)
    wsMrowInc = wsMrowInc + 2

    #assets total per Rich Dad
    ws[column + str(wsMrowInc)] = 'Total Assets Per Rich Dad'
    ws[column + str(wsMrowInc)].font = Font(bold=True)
    ws.merge_cells(column + str(wsMrowAna) + ':' + chr(ord(column)+1) + str(wsMrowAna))
    ws[chr(ord(column)+2) + str(wsMrowInc)] = '=SUM(' + AssetTotal['Total'] + ')'
    ws[chr(ord(column)+2) + str(wsMrowInc)].font = Font(bold=True)
    AssetTotal['TotalPerRichDad'] = chr(ord(column)+2) + str(wsMrowInc)
    wsMrowInc = wsMrowInc + 2
    
    #analysis column
    ws[chr(ord(column)+4) + str(wsMrowAna)] = 'Analysis'
    ws[chr(ord(column)+4) + str(wsMrowAna)].font = Font(bold=True)
    ws.merge_cells(chr(ord(column)+4) + str(wsMrowAna) + ':' + chr(ord(column)+6) + str(wsMrowAna))
    wsMrowAna = wsMrowAna + 1
    #Cash Flow
    ws[chr(ord(column)+4) + str(wsMrowAna)] = 'How much do you keep?'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = 'Cash Flow/Total Income'
    ws[chr(ord(column)+6) + str(wsMrowAna)] = '=SUM(' + NetTotal + '/' + IncTotal['Total'] + ')'
    ws[chr(ord(column)+6) + str(wsMrowAna)].style = 'Percent'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = '***should be increasing'
    ws[chr(ord(column)+5) + str(wsMrowAna)].font = Font(size=8)
    wsMrowAna = wsMrowAna + 4
    ws[chr(ord(column)+4) + str(wsMrowAna)] = 'Does Your Money Work For You?'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = 'Passive+Portfolio/Total Inc'
    ws[chr(ord(column)+6) + str(wsMrowAna)] = '=SUM((' + IncTotal['Portfolio'] + '+' + IncTotal['Pasive'] + ')/' + IncTotal['Total'] + ')'
    ws[chr(ord(column)+6) + str(wsMrowAna)].style = 'Percent'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = '***should be increasing'
    ws[chr(ord(column)+5) + str(wsMrowAna)].font = Font(size=8)
    wsMrowAna = wsMrowAna + 2
    ws[chr(ord(column)+4) + str(wsMrowAna)] = 'How Much Do You Pay In Taxes?'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = 'Income Taxes/Total Income'
    ws[chr(ord(column)+6) + str(wsMrowAna)] = '=SUM(' + ExpTotal['TaxIncome'] + '/' + IncTotal['Total'] + ')'
    ws[chr(ord(column)+6) + str(wsMrowAna)].style = 'Percent'
    wsMrowAna = wsMrowAna + 2
    ws[chr(ord(column)+4) + str(wsMrowAna)] = 'How Much Goes to Housing?'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = 'Housing Expenses/Income'
    ws[chr(ord(column)+6) + str(wsMrowAna)] = '=SUM(' + ExpTotal['Household'] + '/' + IncTotal['Total'] + ')'
    ws[chr(ord(column)+6) + str(wsMrowAna)].style = 'Percent'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = '***keep under 33 percent'
    ws[chr(ord(column)+5) + str(wsMrowAna)].font = Font(size=8)
    wsMrowAna = wsMrowAna + 2
    ws[chr(ord(column)+4) + str(wsMrowAna)] = 'How Much Do You Spend on Depriciating Assets?'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = 'Depreciating Assets Total/Banker Assets'
    ws[chr(ord(column)+6) + str(wsMrowAna)] = '=(' + DepAssetTotal['Total'] + '/' + AssetTotal['TotalPerBanker'] + ')'
    ws[chr(ord(column)+6) + str(wsMrowAna)].style = 'Percent'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = '***keep under 33 percent'
    ws[chr(ord(column)+5) + str(wsMrowAna)].font = Font(size=8)
    wsMrowAna = wsMrowAna + 2
    ws[chr(ord(column)+4) + str(wsMrowAna)] = 'What Is Your Annual Return On Assets?'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = 'Pass+Port/Rich Dad Assets'
    ws[chr(ord(column)+6) + str(wsMrowAna)] = '=((' + IncTotal['Pasive'] + '+' + IncTotal['Portfolio'] + ')*12/' + AssetTotal['TotalPerRichDad'] + ')'
    ws[chr(ord(column)+6) + str(wsMrowAna)].style = 'Percent'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = '***should be increasing'
    ws[chr(ord(column)+5) + str(wsMrowAna)].font = Font(size=8)
    wsMrowAna = wsMrowAna + 2
    ws[chr(ord(column)+4) + str(wsMrowAna)] = 'How Wealthy Are You?'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = 'Rich Dad Assets/Expenses'
    ws[chr(ord(column)+6) + str(wsMrowAna)] = '=(' + AssetTotal['TotalPerRichDad'] + '/' + ExpTotal['Total'] + ')'
    ws[chr(ord(column)+6) + str(wsMrowAna)].style = 'Percent'
    wsMrowAna = wsMrowAna + 1
    ws[chr(ord(column)+5) + str(wsMrowAna)] = '***measured in months'
    ws[chr(ord(column)+5) + str(wsMrowAna)].font = Font(size=8)
    wsMrowAna = wsMrowAna + 2

    # accounts section
    # get a list and value of Current and Savings Accounts from database
    AccTypes = ['CURRENT','SAVINGS']
    AccountsList = getAccountsList(AccTypes)
    print('Getting accounts lists:', AccTypes, AccountsList)
    for a in AccountsList:
        ws[chr(ord(column)+4) + str(wsMrowAna)] = a['accountname']
        wsMrowAna = wsMrowAna + 1
        ws[chr(ord(column)+5) + str(wsMrowAna)] = 'Opening Balance'
        ws[chr(ord(column)+6) + str(wsMrowAna)] = Balance[a['accountname']]['open']['balance']
        wsMrowAna = wsMrowAna + 1
        ws[chr(ord(column)+5) + str(wsMrowAna)] = 'Closing Balance'
        ws[chr(ord(column)+6) + str(wsMrowAna)] = Balance[a['accountname']]['close']['balance']
        wsMrowAna = wsMrowAna + 1
    

    # Labilities section
    # header
    LibilitiesTotal = {}
    libColumn = chr(ord(column)+4)
    ws[libColumn + str(wsMrowLib)] = 'Liabilities'
    ws[libColumn + str(wsMrowLib)].font = Font(bold=True)
    ws.merge_cells(libColumn + str(wsMrowLib) + ':' + chr(ord(libColumn)+2) + str(wsMrowLib))
    wsMHrowLib = wsMrowLib
    wsMrowLib = wsMrowLib + 1
    # get a list and value of Liabilities from database
    AccTypes = ['LOAN']
    Liabilities = database.getAccountNames(AccTypes)
    for l in Liabilities:
        ws[chr(ord(libColumn)+1) + str(wsMrowLib)] = l['accountname']
        LibilitiesTotal[l['accountname']] = chr(ord(libColumn)+2) + str(wsMrowLib)
        ws[LibilitiesTotal[l['accountname']]] = Balance[l['accountname']]['close']['balance']
        wsMrowLib = wsMrowLib + 1
    #liabilities total
    ws[libColumn + str(wsMrowLib)] = 'Liabilities Total'
    ws[libColumn + str(wsMrowLib)].font = Font(bold=True)
    ws.merge_cells(libColumn + str(wsMrowLib) + ':' + chr(ord(libColumn)+1) + str(wsMrowLib))
    ws[chr(ord(libColumn)+2) + str(wsMrowLib)] = '=SUM(' + chr(ord(libColumn)+2) + str(wsMHrowLib + 1) + ':' + chr(ord(libColumn)+2) + str(wsMrowLib - 1) +')'
    ws[chr(ord(libColumn)+2) + str(wsMrowLib)].font = Font(bold=True)
    LibilitiesTotal['Total'] = chr(ord(libColumn)+2) + str(wsMrowLib)
    wsMrowLib = wsMrowLib + 2

    #end of function
    return(ExpTotal, IncTotal)

# Create the Overview table heading row
wsO['A' + str(wsOrow)] = 'Month'
wsO['B' + str(wsOrow)] = 'Actual Expenses'
wsO['C' + str(wsOrow)] = 'Budget expenses'
wsO['D' + str(wsOrow)] = '+/-'
wsO['E' + str(wsOrow)] = 'Actual Income'
wsO['F' + str(wsOrow)] = 'Budget Income'
wsO['G' + str(wsOrow)] = '+/-'
wsO['H' + str(wsOrow)] = 'Actual Balance'
wsO['I' + str(wsOrow)] = 'Budget Balance'
wsO['J' + str(wsOrow)] = '+/-'
wsOHrow = wsOrow
wsOrow = wsOrow + 1

#Create the rows in the overview table
OTotals = {}
for m in months:
    OTotals[m]=create_month(m)
    wsO['A' + str(wsOrow)] = m
    wsO['B' + str(wsOrow)] = '=' + m + '!' + OTotals[m][0]['Total']
    wsO['C' + str(wsOrow)] = '=Budget!' + str(BExpTotal)
    wsO['D' + str(wsOrow)] = '=SUM(B'+ str(wsOrow) + '-C' + str(wsOrow) + ')'
    wsO['E' + str(wsOrow)] = '=' + m +'!' + OTotals[m][1]['Total']
    wsO['F' + str(wsOrow)] = '=Budget!' + str(BIncTotal)
    wsO['G' + str(wsOrow)] = '=SUM(E' + str(wsOrow) + '-F' + str(wsOrow) + ')'
    wsO['H' + str(wsOrow)] = '=SUM(E' + str(wsOrow) + '-B' + str(wsOrow) + ')'
    wsO['I' + str(wsOrow)] = '=SUM(F' + str(wsOrow) + '-C' + str(wsOrow) + ')'
    wsO['J' + str(wsOrow)] = '=SUM(H' + str(wsOrow) + '-I' + str(wsOrow) + ')'
    wsOrow = wsOrow + 1
#DEBUG print(OTotals)

#create the totals row in the overview table
wsO['A' + str(wsOrow)] = 'Totals'
wsO['B' + str(wsOrow)] = '=SUM(B' + str(wsOHrow + 1) + ':B' + str(wsOrow - 1) +')'
wsO['C' + str(wsOrow)] = '=SUM(C' + str(wsOHrow + 1) + ':C' + str(wsOrow - 1) +')'
wsO['D' + str(wsOrow)] = '=SUM(D' + str(wsOHrow + 1) + ':D' + str(wsOrow - 1) +')'
wsO['E' + str(wsOrow)] = '=SUM(E' + str(wsOHrow + 1) + ':E' + str(wsOrow - 1) +')'
wsO['F' + str(wsOrow)] = '=SUM(F' + str(wsOHrow + 1) + ':F' + str(wsOrow - 1) +')'
wsO['G' + str(wsOrow)] = '=SUM(G' + str(wsOHrow + 1) + ':G' + str(wsOrow - 1) +')'
wsO['H' + str(wsOrow)] = '=SUM(H' + str(wsOHrow + 1) + ':H' + str(wsOrow - 1) +')'
wsO['I' + str(wsOrow)] = '=SUM(I' + str(wsOHrow + 1) + ':I' + str(wsOrow - 1) +')'
wsO['J' + str(wsOrow)] = '=SUM(J' + str(wsOHrow + 1) + ':J' + str(wsOrow - 1) +')'

# Save the file
wb.save("Financal Statement.xlsx")
